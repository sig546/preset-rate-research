# -*- coding: utf-8 -*-
"""
生成「预定利率研究值 - 数据验证」Excel。
数据源：data/actuals.json（历史实际公布值）+ data/predictions.json（模型预测值）。
输出：outputs/预定利率研究值_数据验证_<日期>.xlsx
四个工作表：说明、历史实际值、模型预测值、历史vs预测对比时间线。
数值保留 4 位小数，单位标注为 %（百分点差值单位为 BP）。
"""
import json
import os
import calendar
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ACT_FILE = os.path.join(ROOT, "data", "actuals.json")
PRED_FILE = os.path.join(ROOT, "data", "predictions.json")
BOND_FILE = os.path.join(ROOT, "data", "bond_yields.json")
OUT_DIR = os.path.join(ROOT, "outputs")
os.makedirs(OUT_DIR, exist_ok=True)

with open(ACT_FILE, encoding="utf-8") as f:
    act = json.load(f)
with open(PRED_FILE, encoding="utf-8") as f:
    pred = json.load(f)
with open(BOND_FILE, encoding="utf-8") as f:
    _bond_raw = json.load(f)
BOND = _bond_raw.get("data", _bond_raw) if isinstance(_bond_raw, dict) else _bond_raw
_BOND_LAST = BOND[-1]


def _quarter_comp_date(quarter: str) -> str:
    """季度 -> 季末计算日（与 scripts/update_predictions.py 一致）。"""
    y = int(quarter[:4]); q = int(quarter[5:])
    m = {1: 3, 2: 6, 3: 9, 4: 12}[q]
    return "{:04d}-{:02d}-{:02d}".format(y, m, calendar.monthrange(y, m)[1])


def _extend_bond(cutoff: str):
    """未来日期平推（与 update_predictions.py 一致）：晚于最后交易日的工作日按末值填充。"""
    last_date = _BOND_LAST["date"]
    if cutoff <= last_date:
        return list(BOND)
    lg, lc, li = _BOND_LAST["gov_10y"], _BOND_LAST["cdb_10y"], _BOND_LAST["ieb_10y"]
    ext = list(BOND)
    d = date.fromisoformat(last_date) + timedelta(days=1)
    end = date.fromisoformat(cutoff)
    while d <= end:
        if d.weekday() < 5:
            ext.append({"date": d.isoformat(), "gov_10y": lg, "cdb_10y": lc, "ieb_10y": li})
        d += timedelta(days=1)
    return ext


def tax_premium_breakdown(comp_date: str, window: int) -> dict:
    """还原税收溢价(=窗口内每日 max(国开利差,进出口利差) 均值)并拆分子项。

    与 scripts/update_predictions.py 的 compute_spread_ma / compute_ma 完全一致，
    结果可精确复现 predictions.json 的 spread_250 / spread_750。
    """
    data = _extend_bond(comp_date)
    elig = [r for r in data if r["date"] <= comp_date]
    win = elig[-window:]
    n = len(win)
    sum_gov = sum_cdb = sum_ieb = 0.0
    sum_max = 0.0
    cdb_wins = ieb_wins = 0
    for r in win:
        g, c, i = r["gov_10y"], r["cdb_10y"], r["ieb_10y"]
        sum_gov += g; sum_cdb += c; sum_ieb += i
        s_cdb = c - g; s_ieb = i - g
        if s_cdb >= s_ieb:
            cdb_wins += 1
        else:
            ieb_wins += 1
        sum_max += max(s_cdb, s_ieb)
    ma_gov = sum_gov / n; ma_cdb = sum_cdb / n; ma_ieb = sum_ieb / n
    return {
        "comp_date": comp_date,
        "window": window,
        "n": n,
        "start": win[0]["date"],
        "end": win[-1]["date"],
        "ma_gov": ma_gov,          # 国债10Y窗口均值 [A]
        "ma_cdb": ma_cdb,          # 国开债10Y窗口均值 [B]
        "ma_ieb": ma_ieb,          # 进出口债10Y窗口均值 [C]
        "spread_cdb": ma_cdb - ma_gov,   # 国开债利差均值 ① = B-A
        "spread_ieb": ma_ieb - ma_gov,   # 进出口债利差均值 ② = C-A
        "cdb_wins": cdb_wins,
        "ieb_wins": ieb_wins,
        "tax_premium": sum_max / n,       # 税收溢价 = 每日 max(①,②) 窗口均值【最终值】
    }


def fmt_pct(x):
    """格式化利率：去掉多余尾零，但至少保留 2 位小数（如 3.5->3.50，1.7385->1.7385）。"""
    s = "{:.4f}".format(float(x)).rstrip("0").rstrip(".")
    if "." not in s:
        s += ".00"
    elif len(s.split(".")[1]) == 1:
        s += "0"
    return s

# ---------- 样式 ----------
NAVY = "1F3A5F"
BLUE = "2563EB"
LIGHT = "EAF1FB"
GREY = "F5F7FA"
RED = "C0392B"
GREEN = "1E8449"

hdr_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
title_font = Font(name="微软雅黑", size=14, bold=True, color=NAVY)
sub_font = Font(name="微软雅黑", size=9, color="666666")
cell_font = Font(name="微软雅黑", size=10, color="222222")
bold_cell = Font(name="微软雅黑", size=10, bold=True, color=NAVY)

hdr_fill = PatternFill("solid", fgColor=NAVY)
light_fill = PatternFill("solid", fgColor=LIGHT)
grey_fill = PatternFill("solid", fgColor=GREY)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")

thin = Side(style="thin", color="D0D7E2")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border


def apply_border_zebra(ws, r0, r1, ncols):
    for r in range(r0, r1 + 1):
        fill = grey_fill if (r - r0) % 2 == 1 else None
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if cell.font is None or cell.font.name != "微软雅黑":
                cell.font = cell_font
            if fill and cell.fill.fgColor.rgb in ("00000000", None):
                cell.fill = fill


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ============ Sheet 1: 说明 ============
ws0 = wb.active
ws0.title = "说明"
ws0["A1"] = "普通型人身保险产品预定利率研究值 · 数据验证表"
ws0["A1"].font = title_font
ws0.merge_cells("A1:B1")
ws0["A3"] = "生成日期"
ws0["B3"] = date.today().isoformat()
info = [
    ("数据用途", "验证历史实际公布值与模型预测值的准确性，便于对比历史与预测差异"),
    ("历史实际值来源", act.get("description", "中国保险行业协会公布值")),
    ("历史值更新时间", act.get("last_updated", "")),
    ("预测值方法", pred.get("method", "平推法")),
    ("预测值说明", pred.get("description", "")),
    ("预测值更新时间", pred.get("last_updated", "")),
    ("平推法假设", "未来债券收益率、5年期LPR、5年期定存利率均维持当前水平不变"),
    ("当前基础数据", "以下为平推法所采用的“当前水平”，每项均标注实际采用日期与数据来源，可追溯至明确公布日期："),
    ("　· 5年期LPR", "5年期LPR（采用2025年5月20日公布值）={lpr_5y}%　｜来源：中国人民银行授权全国银行间同业拆借中心发布（中国货币网 chinamoney.com.cn）".format(
        lpr_5y=fmt_pct(pred["base_data"]["lpr_5y"]))),
    ("　· 5年期定期存款利率", "5年期定期存款利率（采用2025年5月20日起六大行挂牌值）={dep}%　｜来源：工、农、中、建、交、邮储六大国有商业银行官网挂牌利率".format(
        dep=fmt_pct(pred["base_data"]["deposit_5y"]))),
    ("　· 10年期国债到期收益率", "10年期国债到期收益率（采用2026年7月8日收盘值）={bond}%　｜来源：中国货币网日频行情（东方财富核对）".format(
        bond=fmt_pct(pred["base_data"]["bond_yield_10y"]))),
    ("数值精度", "研究值/收益率均保留小数点后 4 位；单位为百分点（%）；差值单位为 BP（1BP=0.01%）"),
    ("触发规则", "研究值与在售最高值差值连续2个季度≥25BP → 触发预定利率下调/上调"),
]
r = 4
for k, v in info:
    ws0.cell(row=r, column=1, value=k).font = bold_cell
    ws0.cell(row=r, column=1).alignment = left
    ws0.cell(row=r, column=1).fill = light_fill
    ws0.cell(row=r, column=2, value=v).font = cell_font
    ws0.cell(row=r, column=2).alignment = left
    ws0.cell(row=r, column=1).border = border
    ws0.cell(row=r, column=2).border = border
    r += 1
ws0.cell(row=3, column=1).font = bold_cell
ws0.cell(row=3, column=1).fill = light_fill
ws0.cell(row=3, column=1).border = border
ws0.cell(row=3, column=1).alignment = left
ws0.cell(row=3, column=2).font = cell_font
ws0.cell(row=3, column=2).border = border
ws0.cell(row=3, column=2).alignment = left
set_widths(ws0, [22, 88])
ws0.row_dimensions[1].height = 26

# ============ Sheet 2: 历史实际值 ============
ws1 = wb.create_sheet("历史实际值")
ws1["A1"] = "一、预定利率研究值历史实际公布值（来源：中国保险行业协会）"
ws1["A1"].font = title_font
ws1.merge_cells("A1:G1")
ws1["A2"] = "单位：研究值/最高值为百分点(%)，环比变化为BP；数值保留4位小数"
ws1["A2"].font = sub_font
ws1.merge_cells("A2:G2")

headers1 = ["季度", "公布时间", "研究值(%)", "当期在售最高值(%)", "环比变化(BP)", "数据来源", "备注"]
hr = 4
for c, h in enumerate(headers1, start=1):
    ws1.cell(row=hr, column=c, value=h)
style_header(ws1, hr, len(headers1))

r = hr + 1
for a in act["actuals"]:
    ws1.cell(row=r, column=1, value=a["quarter"]).alignment = center
    ws1.cell(row=r, column=2, value=a["announced"]).alignment = center
    cv = ws1.cell(row=r, column=3, value=round(float(a["value"]), 4))
    cv.number_format = "0.0000"
    cv.alignment = right
    cv.font = bold_cell
    mr = ws1.cell(row=r, column=4, value=round(float(a["max_rate"]), 4))
    mr.number_format = "0.0000"
    mr.alignment = right
    chg = a.get("change_bp")
    cc = ws1.cell(row=r, column=5, value=("—" if chg is None else int(chg)))
    cc.alignment = center
    if chg is not None:
        cc.font = Font(name="微软雅黑", size=10, bold=True,
                       color=(RED if chg > 0 else GREEN))  # 涨红跌绿
    ws1.cell(row=r, column=6, value="中国保险行业协会").alignment = center
    ws1.cell(row=r, column=7, value=a.get("note", "")).alignment = left
    r += 1
apply_border_zebra(ws1, hr + 1, r - 1, len(headers1))
set_widths(ws1, [10, 12, 12, 16, 12, 16, 46])
ws1.row_dimensions[1].height = 24

# ============ Sheet 3: 模型预测值 ============
ws2 = wb.create_sheet("模型预测值")
ws2["A1"] = "二、模型对未来各期的预测值（方法：{}）".format(pred.get("method", "平推法"))
ws2["A1"].font = title_font
ws2.merge_cells("A1:K1")
ws2["A2"] = ("单位：利率类为百分点(%)，差值为BP；数值保留4位小数。平推法假设：未来债券收益率、"
            "5年期LPR、5年期定存均维持当前水平不变。")
ws2["A2"].font = sub_font
ws2.merge_cells("A2:K2")

headers2 = ["预测季度", "预测公布时间", "预测研究值(%)", "参考利率(%)", "基础回报(%)",
            "MA250国债(%)", "MA750国债(%)", "税收溢价250(%)", "税收溢价750(%)",
            "距最高值差值(BP)", "是否触发调整"]
hr = 4
for c, h in enumerate(headers2, start=1):
    ws2.cell(row=hr, column=c, value=h)
style_header(ws2, hr, len(headers2))

r = hr + 1
for p in pred["predictions"]:
    ws2.cell(row=r, column=1, value=p["quarter"]).alignment = center
    ws2.cell(row=r, column=2, value=p["announced"]).alignment = center
    pv = ws2.cell(row=r, column=3, value=round(float(p["predicted_value"]), 4))
    pv.number_format = "0.0000"; pv.alignment = right; pv.font = bold_cell
    for col, key in [(4, "ref_rate"), (5, "base_return"), (6, "ma250_gov"),
                     (7, "ma750_gov"), (8, "spread_250"), (9, "spread_750")]:
        cc = ws2.cell(row=r, column=col, value=round(float(p[key]), 4))
        cc.number_format = "0.0000"; cc.alignment = right
    gb = ws2.cell(row=r, column=10, value=round(float(p["gap_bp"]), 1))
    gb.number_format = "0.0"; gb.alignment = center
    tg = ws2.cell(row=r, column=11, value=("是" if p["trigger"] else "否"))
    tg.alignment = center
    tg.font = Font(name="微软雅黑", size=10, bold=True,
                   color=(RED if p["trigger"] else GREEN))
    r += 1
apply_border_zebra(ws2, hr + 1, r - 1, len(headers2))
set_widths(ws2, [11, 13, 13, 11, 11, 12, 12, 13, 13, 15, 13])
ws2.row_dimensions[1].height = 24

# ============ Sheet 4: 历史 vs 预测 对比时间线 ============
ws3 = wb.create_sheet("历史vs预测对比")
ws3["A1"] = "三、历史实际值 vs 模型预测值 · 时间线对比"
ws3["A1"].font = title_font
ws3.merge_cells("A1:E1")
ws3["A2"] = "单位：研究值/最高值为百分点(%)，差值为BP；数值保留4位小数"
ws3["A2"].font = sub_font
ws3.merge_cells("A2:E2")

headers3 = ["季度", "类型", "研究值(%)", "在售最高值(%)", "距最高值差值(BP)"]
hr = 4
for c, h in enumerate(headers3, start=1):
    ws3.cell(row=hr, column=c, value=h)
style_header(ws3, hr, len(headers3))

r = hr + 1
act_fill = PatternFill("solid", fgColor="FDF3E7")   # 历史=浅橙
pred_fill = PatternFill("solid", fgColor="E8F0FE")  # 预测=浅蓝
for a in act["actuals"]:
    gap = round((float(a["value"]) - float(a["max_rate"])) * 100, 1)
    ws3.cell(row=r, column=1, value=a["quarter"]).alignment = center
    t = ws3.cell(row=r, column=2, value="历史实际"); t.alignment = center
    v = ws3.cell(row=r, column=3, value=round(float(a["value"]), 4))
    v.number_format = "0.0000"; v.alignment = right; v.font = bold_cell
    m = ws3.cell(row=r, column=4, value=round(float(a["max_rate"]), 4))
    m.number_format = "0.0000"; m.alignment = right
    g = ws3.cell(row=r, column=5, value=gap); g.number_format = "0.0"; g.alignment = center
    for c in range(1, 6):
        ws3.cell(row=r, column=c).fill = act_fill
        ws3.cell(row=r, column=c).border = border
    r += 1
for p in pred["predictions"]:
    ws3.cell(row=r, column=1, value=p["quarter"]).alignment = center
    t = ws3.cell(row=r, column=2, value="模型预测"); t.alignment = center
    v = ws3.cell(row=r, column=3, value=round(float(p["predicted_value"]), 4))
    v.number_format = "0.0000"; v.alignment = right; v.font = bold_cell
    m = ws3.cell(row=r, column=4, value=round(float(p["max_rate"]), 4))
    m.number_format = "0.0000"; m.alignment = right
    g = ws3.cell(row=r, column=5, value=round(float(p["gap_bp"]), 1))
    g.number_format = "0.0"; g.alignment = center
    for c in range(1, 6):
        ws3.cell(row=r, column=c).fill = pred_fill
        ws3.cell(row=r, column=c).border = border
    r += 1
set_widths(ws3, [12, 12, 13, 16, 16])
ws3.row_dimensions[1].height = 24

# ============ Sheet 5: 税收溢价明细拆分 ============
ws4 = wb.create_sheet("税收溢价明细拆分")
ws4["A1"] = "四、税收溢价（250日 / 750日）明细拆分"
ws4["A1"].font = title_font
ws4.merge_cells("A1:L1")
ws4["A2"] = "单位：收益率/利差均为百分点(%)，保留4位小数；用于核对「模型预测值」表中的税收溢价250 / 税收溢价750"
ws4["A2"].font = sub_font
ws4.merge_cells("A2:L2")

# ---- 计算依据说明块 ----
basis = [
    ("税收溢价定义", "逐个交易日先取「国开债利差」与「进出口债利差」中的较大者，再对最近 N 个交易日求算术平均。"),
    ("　子项① 国开债利差", "国开债利差 = 10年期国开债到期收益率 − 10年期国债到期收益率（逐日）"),
    ("　子项② 进出口债利差", "进出口债利差 = 10年期进出口银行债到期收益率 − 10年期国债到期收益率（逐日）"),
    ("　每日取大者利差", "每日取大者利差 = MAX（国开债利差, 进出口债利差）"),
    ("税收溢价250 / 750", "税收溢价250 = 最近 250 个交易日「每日取大者利差」的均值；税收溢价750 = 最近 750 个交易日的均值。"),
    ("与利差均值的关系", "因逐日取大（非线性），税收溢价 ≥ MAX（国开利差均值, 进出口利差均值），通常略高于「进出口债利差均值」；下表列出两者的取大者天数以便核对差异来源。"),
    ("数据来源", "data/bond_yields.json（10年期国债/国开债/进出口债到期收益率，中国货币网·中债估值，东方财富核对）；2026-07-08 后按平推法维持最后交易日水平。"),
]
r = 4
for k, v in basis:
    kc = ws4.cell(row=r, column=1, value=k); kc.font = bold_cell; kc.alignment = left; kc.fill = light_fill; kc.border = border
    ws4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    vc = ws4.cell(row=r, column=2, value=v); vc.font = cell_font; vc.alignment = left; vc.border = border
    r += 1

headers4 = ["预测季度", "计算截止日", "窗口起始日", "参与交易日数",
            "国债10Y均值A(%)", "国开债10Y均值B(%)", "进出口债10Y均值C(%)",
            "国开债利差均值①=B−A(%)", "进出口债利差均值②=C−A(%)",
            "国开取大者天数", "进出口取大者天数", "税收溢价=每日MAX(①,②)均值(%)"]

quarters = [p["quarter"] for p in pred["predictions"]]


def render_breakdown_table(ws, start_row, window, section_title):
    # 分节标题
    tc = ws.cell(row=start_row, column=1, value=section_title)
    tc.font = Font(name="微软雅黑", size=11, bold=True, color=BLUE)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=12)
    hr = start_row + 1
    for c, h in enumerate(headers4, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers4))
    rr = hr + 1
    for q in quarters:
        b = tax_premium_breakdown(_quarter_comp_date(q), window)
        ws.cell(row=rr, column=1, value=q).alignment = center
        ws.cell(row=rr, column=2, value=b["comp_date"]).alignment = center
        ws.cell(row=rr, column=3, value=b["start"]).alignment = center
        ws.cell(row=rr, column=4, value=b["n"]).alignment = center
        for col, key in [(5, "ma_gov"), (6, "ma_cdb"), (7, "ma_ieb"),
                         (8, "spread_cdb"), (9, "spread_ieb")]:
            cc = ws.cell(row=rr, column=col, value=round(b[key], 4))
            cc.number_format = "0.0000"; cc.alignment = right
        ws.cell(row=rr, column=10, value=b["cdb_wins"]).alignment = center
        ws.cell(row=rr, column=11, value=b["ieb_wins"]).alignment = center
        tp = ws.cell(row=rr, column=12, value=round(b["tax_premium"], 4))
        tp.number_format = "0.0000"; tp.alignment = right; tp.font = bold_cell
        rr += 1
    apply_border_zebra(ws, hr + 1, rr - 1, len(headers4))
    return rr


row_after = render_breakdown_table(ws4, r + 1, 250, "表A · 税收溢价250 明细拆分（最近250个交易日）")
render_breakdown_table(ws4, row_after + 1, 750, "表B · 税收溢价750 明细拆分（最近750个交易日）")
set_widths(ws4, [11, 12, 12, 11, 14, 15, 16, 18, 19, 13, 14, 22])
ws4.row_dimensions[1].height = 24

# 冻结表头
for ws, fr in [(ws1, "A5"), (ws2, "A5"), (ws3, "A5")]:
    ws.freeze_panes = fr

out = os.path.join(OUT_DIR, "预定利率研究值_数据验证_{}.xlsx".format(date.today().isoformat()))
wb.save(out)
print("SAVED:", out)
print("actuals:", len(act["actuals"]), "predictions:", len(pred["predictions"]))
