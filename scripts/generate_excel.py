#!/usr/bin/env python3
"""生成Excel查询表：汇总LPR、定存、研究值、预测值、平安公告等数据"""
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
OUTPUT = PROJECT_ROOT / "预定利率研究值_数据汇总.xlsx"

def load_json(name):
    with open(DATA_DIR / name, "r", encoding="utf-8") as f:
        return json.load(f)

def style_header(ws, row, cols, fill_color="1F4E79"):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=11)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

def auto_width(ws, cols):
    for c in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(max_len + 4, 30)

def main():
    wb = Workbook()

    # ── Sheet 1: 预定利率研究值 ──
    ws1 = wb.active
    ws1.title = "预定利率研究值"
    headers = ["季度", "公布时间", "研究值(%)", "环比(BP)", "最高值(%)", "差值(BP)", "类型", "说明"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers))

    actuals = load_json("actuals.json")["actuals"]
    preds = load_json("predictions.json")["predictions"]

    row = 2
    prev_val = None
    for a in actuals:
        ch = round((a["value"] - prev_val) * 100, 1) if prev_val is not None else None
        prev_val = a["value"]
        gap = round((a["max_rate"] - a["value"]) * 100, 1)
        ws1.cell(row=row, column=1, value=a["quarter"])
        ws1.cell(row=row, column=2, value=a["announced"])
        ws1.cell(row=row, column=3, value=a["value"])
        ws1.cell(row=row, column=4, value=ch)
        ws1.cell(row=row, column=5, value=a["max_rate"])
        ws1.cell(row=row, column=6, value=gap)
        ws1.cell(row=row, column=7, value="实际值")
        ws1.cell(row=row, column=8, value=a.get("note", ""))
        row += 1

    for p in preds:
        ch = round((p["predicted_value"] - prev_val) * 100, 1)
        prev_val = p["predicted_value"]
        ws1.cell(row=row, column=1, value=p["quarter"])
        ws1.cell(row=row, column=2, value=p["announced"])
        ws1.cell(row=row, column=3, value=round(p["predicted_value"], 4))
        ws1.cell(row=row, column=4, value=ch)
        ws1.cell(row=row, column=5, value=p["max_rate"])
        ws1.cell(row=row, column=6, value=round(p["gap_bp"], 1))
        ws1.cell(row=row, column=7, value="预测值")
        ws1.cell(row=row, column=8, value="平推法")
        row += 1
    auto_width(ws1, len(headers))

    # ── Sheet 2: 5Y LPR 历史 ──
    ws2 = wb.create_sheet("5Y LPR历史")
    headers2 = ["调整日期", "利率(%)", "变动幅度(BP)", "说明"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2), "2E75B6")

    lpr = load_json("lpr_history.json")["history"]
    prev_r = None
    for i, e in enumerate(lpr):
        r = e["rate"]
        ch = round((r - prev_r) * 100, 1) if prev_r is not None else None
        prev_r = r
        ws2.cell(row=i+2, column=1, value=e["date"])
        ws2.cell(row=i+2, column=2, value=r)
        ws2.cell(row=i+2, column=3, value=ch)
        ws2.cell(row=i+2, column=4, value=e.get("note", ""))
    auto_width(ws2, len(headers2))

    # ── Sheet 3: 5Y 定存利率历史 ──
    ws3 = wb.create_sheet("5Y定存利率历史")
    headers3 = ["调整日期", "利率(%)", "变动幅度(BP)", "来源"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, len(headers3), "548235")

    deposit = load_json("deposit_rates.json")["history"]
    prev_r = None
    for i, e in enumerate(deposit):
        r = e["rate"]
        ch = round((r - prev_r) * 100, 1) if prev_r is not None else None
        prev_r = r
        ws3.cell(row=i+2, column=1, value=e["date"])
        ws3.cell(row=i+2, column=2, value=r)
        ws3.cell(row=i+2, column=3, value=ch)
        ws3.cell(row=i+2, column=4, value=e.get("source", ""))
    auto_width(ws3, len(headers3))

    # ── Sheet 4: 平安人寿公告 ──
    ws4 = wb.create_sheet("平安人寿公告")
    headers4 = ["公告日期", "产品类型", "最高值(%)", "生效时间", "来源"]
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header(ws4, 1, len(headers4), "BF4F51")

    pingan = load_json("pingan_announcements.json")["announcements"]
    for i, a in enumerate(pingan):
        ws4.cell(row=i+2, column=1, value=a["date"])
        ws4.cell(row=i+2, column=2, value=a.get("product_type", "普通型"))
        ws4.cell(row=i+2, column=3, value=a["max_rate"])
        ws4.cell(row=i+2, column=4, value=a.get("effective_date", ""))
        ws4.cell(row=i+2, column=5, value=a.get("source_url", ""))
    auto_width(ws4, len(headers4))

    # ── Sheet 5: 预测详情 ──
    ws5 = wb.create_sheet("预测详情")
    headers5 = ["季度", "公布时间", "预测值(%)", "参考利率(%)", "基础回报(%)", "MA250(%)", "MA750(%)", "利差250(%)", "利差750(%)", "差值(BP)", "是否触发"]
    for c, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=c, value=h)
    style_header(ws5, 1, len(headers5), "ED7D31")

    for i, p in enumerate(preds):
        ws5.cell(row=i+2, column=1, value=p["quarter"])
        ws5.cell(row=i+2, column=2, value=p["announced"])
        ws5.cell(row=i+2, column=3, value=round(p["predicted_value"], 4))
        ws5.cell(row=i+2, column=4, value=round(p.get("ref_rate", 0), 4))
        ws5.cell(row=i+2, column=5, value=round(p.get("base_return", 0), 4))
        ws5.cell(row=i+2, column=6, value=round(p.get("ma250_gov", 0), 4))
        ws5.cell(row=i+2, column=7, value=round(p.get("ma750_gov", 0), 4))
        ws5.cell(row=i+2, column=8, value=round(p.get("spread_250", 0), 4))
        ws5.cell(row=i+2, column=9, value=round(p.get("spread_750", 0), 4))
        ws5.cell(row=i+2, column=10, value=round(p.get("gap_bp", 0), 1))
        ws5.cell(row=i+2, column=11, value="是" if p.get("trigger") else "否")
    auto_width(ws5, len(headers5))

    # Save
    wb.save(OUTPUT)
    print(f"Excel 已生成: {OUTPUT}")
    print(f"  5 个 Sheet: 预定利率研究值 / LPR / 定存 / 平安公告 / 预测详情")

if __name__ == "__main__":
    main()
