import openpyxl

file_path = "C:/Users/marsh/Downloads/预定利率研究值-使用版.xlsx"
wb_data = openpyxl.load_workbook(file_path, data_only=True)
ws = wb_data['预定利率']

print("=== B12到F18数据（系数调节部分）===")
for row in range(12, 19):
    b_val = ws.cell(row=row, column=2).value
    c_val = ws.cell(row=row, column=3).value
    d_val = ws.cell(row=row, column=4).value
    e_val = ws.cell(row=row, column=5).value
    f_val = ws.cell(row=row, column=6).value
    print(f"Row {row}: B={b_val}, C={c_val}, D={d_val}, E={e_val}, F={f_val}")

print("\n=== B3到B7标签 ===")
for row in range(3, 8):
    b_val = ws.cell(row=row, column=2).value
    c_val = ws.cell(row=row, column=3).value
    print(f"Row {row}: B={b_val}, C={c_val}")

print("\n=== H5到H12标签 ===")
for row in range(5, 13):
    for col in range(1, 10):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            print(f"{openpyxl.utils.get_column_letter(col)}{row} = {val}")
