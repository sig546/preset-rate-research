import pandas as pd
import openpyxl

file_path = "C:/Users/marsh/Downloads/预定利率研究值-使用版.xlsx"
wb_data = openpyxl.load_workbook(file_path, data_only=True)

# 检查各sheet的数据范围
for sheet_name in ['国债-到期', '国开债-到期', '进出口-到期', '5年期LPR', '5年期定期存款利率']:
    ws = wb_data[sheet_name]
    print(f"\n=== {sheet_name} ===")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    
    # 显示前5行和最后5行
    print("前5行:")
    for row in range(1, min(6, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(6, ws.max_column + 1)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                row_data.append(f"{openpyxl.utils.get_column_letter(col)}{row}={val}")
        if row_data:
            print(row_data)
    
    print("最后5行:")
    for row in range(max(1, ws.max_row - 4), ws.max_row + 1):
        row_data = []
        for col in range(1, min(6, ws.max_column + 1)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                row_data.append(f"{openpyxl.utils.get_column_letter(col)}{row}={val}")
        if row_data:
            print(row_data)

# 检查国债-到期sheet的完整数据
print("\n=== 国债-到期 详细结构 ===")
ws = wb_data['国债-到期']
for row in range(1, 10):
    row_data = []
    for col in range(1, 10):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            row_data.append(f"{openpyxl.utils.get_column_letter(col)}{row}={val}")
    if row_data:
        print(row_data)

# 检查是否有2025年数据
print("\n=== 检查2025年数据 ===")
ws = wb_data['国债-到期']
found_2025 = False
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and isinstance(val, str) and '2025' in str(val):
        print(f"Row {row}: {val}")
        found_2025 = True
        break
    elif val and hasattr(val, 'year') and val.year == 2025:
        print(f"Row {row}: {val}")
        found_2025 = True
        break

if not found_2025:
    print("未找到2025年数据")
