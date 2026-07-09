import openpyxl

file_path = "C:/Users/marsh/Downloads/预定利率研究值-使用版.xlsx"
wb_data = openpyxl.load_workbook(file_path, data_only=True)

# 检查5年期LPR数据
print("=== 5年期LPR数据 ===")
ws = wb_data['5年期LPR']
for row in range(1, 23):
    b_val = ws.cell(row=row, column=2).value
    c_val = ws.cell(row=row, column=3).value
    if b_val is not None or c_val is not None:
        print(f"Row {row}: B={b_val}, C={c_val}")

# 检查5年期定期存款利率
print("\n=== 5年期定期存款利率 ===")
ws2 = wb_data['5年期定期存款利率']
for row in range(1, 10):
    row_data = []
    for col in range(1, 10):
        val = ws2.cell(row=row, column=col).value
        if val is not None:
            row_data.append(f"{openpyxl.utils.get_column_letter(col)}{row}={val}")
    if row_data:
        print(row_data)

# 找2024-12-31的存款利率
print("\n=== 2024年存款利率 ===")
for row in range(1, 61):
    b_val = ws2.cell(row=row, column=2).value
    if b_val and '2024-12' in str(b_val):
        row_data = []
        for col in range(1, 10):
            val = ws2.cell(row=row, column=col).value
            if val is not None:
                row_data.append(f"{openpyxl.utils.get_column_letter(col)}{row}={val}")
        print(row_data)

# 检查预定利率中L列的完整数据
print("\n=== 预定利率 L列数据 ===")
ws3 = wb_data['预定利率']
for row in range(5, 13):
    l_val = ws3.cell(row=row, column=12).value
    i_val = ws3.cell(row=row, column=9).value
    j_val = ws3.cell(row=row, column=10).value
    k_val = ws3.cell(row=row, column=11).value
    print(f"Row {row}: I={i_val}, J={j_val}, K={k_val}, L={l_val}")
