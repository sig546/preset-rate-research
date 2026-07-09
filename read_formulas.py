import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np

file_path = "C:/Users/marsh/Downloads/预定利率研究值-使用版.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=False)
print("所有sheet名称:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    
    for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=False):
        row_data = []
        for cell in row:
            if cell.value is not None:
                col_letter = get_column_letter(cell.column)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    row_data.append(f"{col_letter}{cell.row}={cell.value}")
                else:
                    row_data.append(f"{col_letter}{cell.row}={cell.value}")
        if row_data:
            print(row_data)
    
    # Show more rows to find formulas
    print(f"\n--- 寻找公式 (更多行) ---")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                col_letter = get_column_letter(cell.column)
                print(f"{col_letter}{cell.row}: {cell.value}")
