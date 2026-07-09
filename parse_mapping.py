import pandas as pd
import openpyxl

file_path = "C:/Users/marsh/Downloads/预定利率研究值-使用版.xlsx"
wb_data = openpyxl.load_workbook(file_path, data_only=True)

# 解析"预定利率"sheet中N列和O-X列的对应关系
ws = wb_data['预定利率']

print("=== N列（搜索键）与O-X列数据映射 ===")
for row in range(12, 19):
    n_val = ws.cell(row=row, column=14).value  # N=14
    o_val = ws.cell(row=row, column=15).value  # O=15
    p_val = ws.cell(row=row, column=16).value  # P=16
    q_val = ws.cell(row=row, column=17).value  # Q=17
    r_val = ws.cell(row=row, column=18).value  # R=18
    s_val = ws.cell(row=row, column=19).value  # S=19
    t_val = ws.cell(row=row, column=20).value  # T=20
    u_val = ws.cell(row=row, column=21).value  # U=21
    v_val = ws.cell(row=row, column=22).value  # V=22
    w_val = ws.cell(row=row, column=23).value  # W=23
    x_val = ws.cell(row=row, column=24).value  # X=24
    print(f"Row {row}: N={n_val}, O={o_val}, P={p_val}, Q={q_val}, R={r_val}, S={s_val}, T={t_val}, U={u_val}, V={v_val}, W={w_val}, X={x_val}")

# 检查国债sheet中的列标签
print("\n=== 国债-到期 Sheet 列标签 (Row 5) ===")
ws_gz = wb_data['国债-到期']
for col in range(1, 20):
    val = ws_gz.cell(row=5, column=col).value
    if val is not None:
        print(f"{openpyxl.utils.get_column_letter(col)}5 = {val}")

# 检查国债sheet中Row 6 (250MA)的数据
print("\n=== 国债-到期 250MA (Row 6) ===")
for col in range(1, 20):
    val = ws_gz.cell(row=6, column=col).value
    if val is not None:
        print(f"{openpyxl.utils.get_column_letter(col)}6 = {val}")

# 检查国债sheet中Row 2 (750MA)的数据
print("\n=== 国债-到期 750MA (Row 2) ===")
for col in range(1, 20):
    val = ws_gz.cell(row=2, column=col).value
    if val is not None:
        print(f"{openpyxl.utils.get_column_letter(col)}2 = {val}")

# 检查input> sheet的offset对应表
print("\n=== input> sheet offset对应表 ===")
ws_input = wb_data['input>']
for row in range(1, 10):
    c_val = ws_input.cell(row=row, column=3).value
    d_val = ws_input.cell(row=row, column=4).value
    if c_val is not None or d_val is not None:
        print(f"Row {row}: C={c_val}, D={d_val}")

# 检查预定利率sheet中H列（offset）
print("\n=== 预定利率 H列（offset）===")
for row in range(5, 13):
    h_val = ws.cell(row=row, column=8).value
    if h_val is not None:
        print(f"H{row} = {h_val}")
