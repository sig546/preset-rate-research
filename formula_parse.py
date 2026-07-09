import pandas as pd
import openpyxl

file_path = "C:/Users/marsh/Downloads/预定利率研究值-使用版.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

# 解析"预定利率"sheet中O到X列的公式（基础回报水平部分）
ws = wb['预定利率']

print("=== 基础回报水平相关公式（O-X列）===")
for row in range(7, 19):
    for col in range(15, 25):  # O=15 to X=24
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(col)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"{col_letter}{row}: {cell.value}")

print("\n=== 参考利率相关公式（H-L列）===")
for row in range(5, 13):
    for col in range(8, 13):  # H=8 to L=12
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(col)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"{col_letter}{row}: {cell.value}")

print("\n=== 最终计算公式（B-C列）===")
for row in range(3, 7):
    for col in range(2, 4):  # B=2 to C=3
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(col)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"{col_letter}{row}: {cell.value}")

# 读取数据来验证
print("\n=== 读取验证数据 ===")
ws_data = openpyxl.load_workbook(file_path, data_only=True)['预定利率']
print(f"C1 (时点): {ws_data['C1'].value}")
print(f"C3 (参考利率): {ws_data['C3'].value}")
print(f"C4 (基础回报水平): {ws_data['C4'].value}")
print(f"C5 (系数调节前): {ws_data['C5'].value}")
print(f"C6 (系数调节): {ws_data['C6'].value}")
print(f"C7 (最终预定利率): {ws_data['C7'].value}")

# 读取关键中间变量
print(f"\nL5 (参考利率计算): {ws_data['L5'].value}")
print(f"P5 (基础回报水平): {ws_data['P5'].value}")
print(f"P7 (250MA国债+税): {ws_data['P7'].value}")
print(f"U7 (750MA国债+税): {ws_data['U7'].value}")
print(f"P8 (250MA累加): {ws_data['P8'].value}")
print(f"U8 (750MA累加): {ws_data['U8'].value}")
print(f"P9 (250MA最终): {ws_data['P9'].value}")
print(f"U9 (750MA最终): {ws_data['U9'].value}")

# 读取5年期LPR和存款利率
print(f"\nJ7 (5Y LPR): {ws_data['J7'].value}")
print(f"K7 (5Y 存款利率): {ws_data['K7'].value}")
print(f"L7 (LPR+存款): {ws_data['L7'].value}")

# 读取国债相关
print(f"\nO15 (国债): {ws_data['O15'].value}")
print(f"P15 (国开债): {ws_data['P15'].value}")
print(f"Q15 (进出口): {ws_data['Q15'].value}")
print(f"R15 (国开-国债): {ws_data['R15'].value}")
print(f"S15 (进出口-国债): {ws_data['S15'].value}")
print(f"T15 (MAX利差): {ws_data['T15'].value}")
